from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from dateutil import parser
from datetime import *
import pytz
import constants
import api

now = datetime.now(pytz.timezone(constants.TIME_ZONE))
yesterday = now - timedelta(1)

def loadWorkbook(self):
    print "loading workbook..."
    wb = load_workbook('./pat.xlsx')
    sheet_names = wb.get_sheet_names()
    today_date_sheet_exists = False
    # Go through all of sheet names, set flag to true if one of the sheet names is today's month and year
    for name in sheet_names:
        if (name != "template"):
            sheet_date = parser.parse(str(name))
            if now.year == sheet_date.year and now.month == sheet_date.month:
                today_date_sheet_exists = True
    active_sheet = ""
    today_sheet_name = str(now.strftime('%Y')) + "-" + str(now.strftime('%m'))
    self.today_date = str(now.strftime('%-m')) + "/" + str(now.strftime('%-d')) + "/" + str(now.strftime('%y'))
    self.yesterday = str(yesterday.strftime('%-m')) + "/" + str(yesterday.strftime('%-d')) + "/" + str(yesterday.strftime('%y'))
    if (today_date_sheet_exists):
        print "editing existing sheet.."
        active_sheet = wb.get_sheet_by_name(today_sheet_name)
    else:
        print "creating new sheet.."
        template_sheet = wb.get_sheet_by_name("template")
        active_sheet = wb.copy_worksheet(template_sheet)
        active_sheet.title = str(now.strftime('%Y')) + "-" + str(now.strftime('%m'))
        # Create data validation for outage types
        dv = DataValidation(type="list", formula1='"No Outage,CaaS Unplanned,CaaS Planned,BSP Infrastructure Unplanned,BSP Infrastructure Planned,IAE Unplanned,IAE Planned,Other"', allow_blank=True)
        active_sheet.add_data_validation(dv)
        tmp_col = constants.init_outage_type_col
        while tmp_col < constants.last_index_col:
            tmp_row = constants.init_outage_type_row
            while tmp_row < constants.last_index_row:
                dv.add(active_sheet.cell(row=tmp_row,column=tmp_col))
                tmp_row += 1
            tmp_col += 7
        # Create dates on left side
        day = 1
        tmp_row = constants.init_dates_row
        while tmp_row <= constants.last_index_row:
            date_to_write = str(now.strftime('%-m')) + "/" + str(day) + "/" + str(now.strftime('%y'))
            active_sheet.cell(row=tmp_row,column=constants.dates_col).value = date_to_write
            day += 1
            tmp_row += 14
        print "finished creating new sheet"

    #writeResolvedIncidentsToExcel(self,wb,active_sheet)
    scanForOpenIncidents(self,wb,active_sheet)
    writeIncidentsToExcel(self,wb,active_sheet,self.resolved_incidents_array)
    writeIncidentsToExcel(self,wb,active_sheet,self.open_incidents_array)
    saveWorkbook(self,wb)

def writeIncidentsToExcel(self,wb,active_sheet,incidents_array):
    wb.active = wb.index(active_sheet)
    tmp_col = constants.init_http_check_col
    while tmp_col <= constants.last_index_col:
        check_name = active_sheet.cell(row=constants.http_check_row,column=tmp_col).value
        for incident in incidents_array:
            titleExists = False
            if incident.title == "check-api-umbrella-status-non-prod" or incident.title == "check_docker dtr-":
                if check_name in incident.title:
                    titleExists = True
            elif check_name == incident.title:
                titleExists = True
            tmp_row = constants.init_dates_row
            while tmp_row <= constants.last_index_row:
                date = active_sheet.cell(row=tmp_row,column=constants.dates_col).value
                if date == self.today_date:
                    if titleExists:
                        row_increment = 1
                        while row_increment <= 11:
                            if not active_sheet.cell(row=tmp_row + row_increment,column=tmp_col + 1).value and not active_sheet.cell(row=tmp_row + row_increment,column=tmp_col + 2).value:
                                # RESOLVED INCIDENTS
                                if incidents_array is self.resolved_incidents_array:
                                    print "writing resolved " + incident.title + " incident.."
                                    # Write start time
                                    active_sheet.cell(row=tmp_row + row_increment,column=tmp_col + 1).value = incident.created_at
                                    # Write end time
                                    active_sheet.cell(row=tmp_row + row_increment,column=tmp_col + 2).value = incident.resolved_time
                                    break
                                # OPEN INCIDENTS
                                else:
                                    print "writing open " + incident.title + " incident.."
                                    # Write start time
                                    active_sheet.cell(row=tmp_row + row_increment,column=tmp_col + 1).value = incident.created_at
                                    # Write end time
                                    active_sheet.cell(row=tmp_row + row_increment,column=tmp_col + 2).value = "*"
                                    # Write incident ID
                                    active_sheet.cell(row=tmp_row + row_increment,column=tmp_col + 5).value = incident.id
                                    break
                            row_increment += 1
                tmp_row += 14
        # write no outage if there were no incidents
        if incidents_array is self.resolved_incidents_array:
            tmp_row = constants.init_dates_row
            while tmp_row <= constants.last_index_row:
                date = active_sheet.cell(row=tmp_row,column=constants.dates_col).value
                if date == self.today_date:
                    if not active_sheet.cell(row=tmp_row + 2,column=tmp_col + 1).value:
                        active_sheet.cell(row=tmp_row + 2,column=tmp_col + 4).value = "No Outage"
                tmp_row += 14
        tmp_col += 7

def scanForOpenIncidents(self,wb,active_sheet):
    if now.day == 1:
        active_sheet = wb.get_sheet_by_name(str(yesterday.strftime('%Y')) + "-" + str(yesterday.strftime('%m')))
    tmp_col = constants.init_outage_end_time
    while tmp_col <= constants.last_index_col:
        tmp_row = constants.init_dates_row
        while tmp_row <= constants.last_index_row:
            date = active_sheet.cell(row=tmp_row,column=constants.dates_col).value
            if date == self.yesterday:
                row_increment = 2
                while row_increment <= 11:
                    if active_sheet.cell(row=tmp_row + row_increment,column=tmp_col).value == "*":
                        #use incident id to find resolved time
                        incident_id = active_sheet.cell(row=tmp_row + row_increment,column=tmp_col + 3).value
                        incident_json = api.getIncidents(self,incident_id)
                        resolved_time_string = ""
                        if incident_json['incident']['status'] == "resolved":
                            resolved_time = parser.parse(incident_json['incident']['last_status_change_at'])
                            resolved_time_string = resolved_time.strftime('%I:%M %p')
                        else:
                            resolved_time_string = "*"

                        #change star to 11:59 pm
                        active_sheet.cell(row=tmp_row + row_increment,column=tmp_col).value ="11:59 PM"
                        #go to today, write 12:01 AM on created at
                        if now.day != 1:
                            active_sheet.cell(row=tmp_row + 16,column=tmp_col - 1).value = "12:01 AM"
                            active_sheet.cell(row=tmp_row + 16,column=tmp_col).value = resolved_time_string
                        #if it's the first, need to change sheets and reset row
                        else:
                            active_sheet = wb.get_sheet_by_name(str(now.strftime('%Y')) + "-" + str(now.strftime('%m')))
                            active_sheet.cell(row=constants.init_outage_type_row,column=tmp_col - 1).value = "12:01 AM"
                            active_sheet.cell(row=constants.init_outage_type_row,column=tmp_col).value = resolved_time_string
                            active_sheet = wb.get_sheet_by_name(str(yesterday.strftime('%Y')) + "-" + str(yesterday.strftime('%m')))
                    row_increment += 1
            tmp_row += 14
        tmp_col += 7


def saveWorkbook(self,wb):
    print "saving workbook.."
    wb.save('./pat.xlsx')
    print "finished!"
